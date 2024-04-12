import errno
import typing

import openpyxl
import pandas
from openpyxl import Workbook


def read_excel(filename: str) -> pandas.DataFrame:
    data = pandas.read_excel(filename)
    return data


def write_excel(filename: str, arr: list, x: float, step: int, deltaZ: float, count=1) -> int:
    try:
        book = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(errno)
        book = Workbook()

    sheet = book.active

    row_count = count

    if step == 0:
        sheet.cell(row=1, column=1).value = \
            f"Kарта распространения вещества: x = 0, step = 0, deltaZ = deltaY = {round(deltaZ, 2)}"
    else:
        sheet.cell(row=row_count, column=1).value = f'x = {x}, step = {step}'

    row_count += 1
    for row in arr:
        sheet.append(row)
        row_count += 1

    book.save(filename)
    return row_count
